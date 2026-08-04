"""
Failed Charging Session Analyzer
---------------------------------
Reads a Charging Sessions export (CSV or Excel), flags every session whose
Duration is below a cutoff (default: 2 minutes) as a "Failed" session, and
produces a ready-to-share Excel report with:

    1. All Sessions      - original data + Duration_Seconds + Is_Failed columns
    2. Failed Sessions   - only the failed rows
    3. Monthly Summary   - total vs failed sessions per month, failed %, 
                            total & average CutOff Amount tied to failed sessions
    4. Weekly Summary    - same breakdown per week
    5. CutOff Summary    - how failed sessions are distributed across 
                            CutOff Amount bands

HOW TO RUN
----------
1. Put this file and your exported file (e.g. ChargingSessions.csv or .xlsx)
   in the same folder.
2. Run from a terminal:

       python failed_session_analyzer.py ChargingSessions.csv

   Optional: change the failure cutoff (in minutes) and output name:

       python failed_session_analyzer.py ChargingSessions.csv --minutes 2 --output report.xlsx

3. Open the generated Excel file (default: Failed_Session_Report.xlsx).

You can also import `analyze()` into another script/notebook and call it
directly with a DataFrame or file path - that's the "automation" hook if you
want to later trigger this from a Flask upload endpoint, a scheduled job, etc.
"""

import argparse
import re
import sys
from pathlib import Path

import pandas as pd
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.enums import TA_LEFT


# --------------------------------------------------------------------------
# 1. Helpers
# --------------------------------------------------------------------------

def parse_duration_to_seconds(value):
    """
    Converts values like '3m', '47s', '1h 5m', '0', '' or NaN into seconds.
    Blank/NaN duration is treated as 0 seconds (i.e. the session never
    really started - a failure).
    """
    if pd.isna(value):
        return 0

    value = str(value).strip()
    if value in ("", "0"):
        return 0

    total_seconds = 0
    # Matches any combination like "1h", "5m", "30s" appearing in the string
    for amount, unit in re.findall(r"(\d+)\s*([hms])", value):
        amount = int(amount)
        if unit == "h":
            total_seconds += amount * 3600
        elif unit == "m":
            total_seconds += amount * 60
        elif unit == "s":
            total_seconds += amount

    return total_seconds


def load_file(file_path):
    """Reads .csv, .xlsx or .xls into a DataFrame."""
    file_path = Path(file_path)
    if file_path.suffix.lower() == ".csv":
        df = pd.read_csv(file_path)
    elif file_path.suffix.lower() in (".xlsx", ".xls"):
        df = pd.read_excel(file_path)
    else:
        raise ValueError(f"Unsupported file type: {file_path.suffix}")
    return df


# --------------------------------------------------------------------------
# 2b. One-page PDF report (styled like a formal statement)
# --------------------------------------------------------------------------

GOLD = colors.HexColor("#C9BE86")   # header band colour, similar to the reference report
DARK = colors.HexColor("#222222")

def generate_pdf_report(df, monthly, weekly, cutoff_summary, cutoff_minutes,
                         pdf_path="Failed_Session_Report.pdf",
                         organisation="Charging Sessions", period_from=None, period_to=None):
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("TitleBig", parent=styles["Title"], fontSize=17,
                                  alignment=TA_LEFT, spaceAfter=1, leading=20)
    sub_style = ParagraphStyle("Sub", parent=styles["Normal"], fontSize=9,
                                textColor=DARK, spaceAfter=2)
    h2_style = ParagraphStyle("H2", parent=styles["Heading2"], fontSize=10.5,
                               spaceBefore=6, spaceAfter=3)
    small = ParagraphStyle("Small", parent=styles["Normal"], fontSize=7, textColor=colors.grey)

    doc = SimpleDocTemplate(pdf_path, pagesize=A4,
                             topMargin=12 * mm, bottomMargin=22 * mm,
                             leftMargin=14 * mm, rightMargin=14 * mm)
    story = []

    story.append(Paragraph("Charging Session Success Report", title_style))

    if period_from is not None and period_to is not None:
        story.append(Paragraph(
            f"Period: {period_from:%d %b %Y} to {period_to:%d %b %Y} &nbsp;&nbsp;|&nbsp;&nbsp; "
            f"Failure cutoff: sessions under {cutoff_minutes} minute(s)",
            sub_style))

    # --- Organisation / Charge Point / Connector identification ----------
    def _unique_join(series, limit=3):
        vals = [str(v) for v in series.dropna().unique()]
        if not vals:
            return "N/A"
        if len(vals) > limit:
            return ", ".join(vals[:limit]) + f" (+{len(vals) - limit} more)"
        return ", ".join(vals)

    org_text = _unique_join(df["Organisation"]) if "Organisation" in df.columns else "N/A"
    cp_text = _unique_join(df["Charge Point Id"]) if "Charge Point Id" in df.columns else "N/A"
    connector_text = _unique_join(df["Connector"]) if "Connector" in df.columns else "N/A"

    story.append(Paragraph(f"<b>Organisation:</b> {org_text}", sub_style))
    story.append(Paragraph(f"<b>Charge Point Id:</b> {cp_text}", sub_style))
    story.append(Paragraph(f"<b>Connector:</b> {connector_text}", sub_style))

    # --- Summary block ---------------------------------------------------
    total = len(df)
    failed = int(df["Is_Failed"].sum())
    success = total - failed
    failed_pct = round(failed / total * 100, 1) if total else 0
    success_pct = round(100 - failed_pct, 1) if total else 0
    total_cutoff_failed = df.loc[df["Is_Failed"], "CutOff Amount_Num"].sum()

    story.append(Paragraph("Summary", h2_style))
    summary_data = [
        ["Total Sessions", "Failed Sessions", "Success Sessions", "Failed %", "Success %", "CutOff Amt (Failed)"],
        [str(total), str(failed), str(success), f"{failed_pct}%", f"{success_pct}%", f"{total_cutoff_failed:,.2f}"],
    ]
    summary_table = Table(summary_data, colWidths=[30 * mm] * 6)
    summary_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), GOLD),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke]),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    story.append(summary_table)

    # --- Monthly summary ---------------------------------------------------
    story.append(Paragraph("Monthly Summary", h2_style))
    m_header = ["Month", "Total", "Failed", "Failed %", "Success %", "CutOff Total", "CutOff Avg"]
    m_rows = [m_header] + [
        [row["Month"], int(row["Total_Sessions"]), int(row["Failed_Sessions"]),
         f"{row['Failed_%']}%", f"{row['Success_%']}%",
         f"{row['Total_CutOff_Amount']:,.2f}", f"{row['Avg_CutOff_Amount']:,.2f}"]
        for _, row in monthly.iterrows()
    ]
    monthly_table = Table(m_rows, colWidths=[24 * mm, 20 * mm, 20 * mm, 20 * mm, 20 * mm, 26 * mm, 26 * mm])
    monthly_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), GOLD),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7.5),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.whitesmoke]),
        ("TOPPADDING", (0, 0), (-1, -1), 2.5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2.5),
    ]))
    story.append(monthly_table)

    # --- Weekly summary ---------------------------------------------------
    story.append(Paragraph("Weekly Summary", h2_style))
    w_header = ["Week", "Total", "Failed", "Failed %", "Success %", "CutOff Total", "CutOff Avg"]
    w_rows = [w_header] + [
        [row["Week"], int(row["Total_Sessions"]), int(row["Failed_Sessions"]),
         f"{row['Failed_%']}%", f"{row['Success_%']}%",
         f"{row['Total_CutOff_Amount']:,.2f}", f"{row['Avg_CutOff_Amount']:,.2f}"]
        for _, row in weekly.iterrows()
    ]
    weekly_table = Table(w_rows, colWidths=[24 * mm, 20 * mm, 20 * mm, 20 * mm, 20 * mm, 26 * mm, 26 * mm],
                          repeatRows=1)
    weekly_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), GOLD),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 6.3),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.whitesmoke]),
        ("TOPPADDING", (0, 0), (-1, -1), 1.6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 1.6),
    ]))
    story.append(weekly_table)

    story.append(Spacer(1, 10))

    def _footer(canvas, doc):
        canvas.saveState()
        width, height = A4
        footer_text = ("Sector No. 5, Auric City, Plot 36, Shendra MIDC, Kubhephal, Maharashtra 431154 "
                        "| GSTIN : 27AAFCI3802C1Z1 | CIN : U31900PN2019PTC186234 "
                        "| Contact Us : support@goegonetwork.com")
        # thin gold divider line above the footer text
        canvas.setFillColor(GOLD)
        canvas.rect(0, 16 * mm, width, 1.2, fill=1, stroke=0)
        # footer text
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(colors.grey)
        canvas.drawCentredString(width / 2, 11 * mm, footer_text)
        # dark bottom bar
        canvas.setFillColor(colors.HexColor("#1a1a1a"))
        canvas.rect(0, 0, width, 5 * mm, fill=1, stroke=0)
        canvas.restoreState()

    doc.build(story, onFirstPage=_footer, onLaterPages=_footer)
    return pdf_path


def analyze(file_path, output_path="Failed_Session_Report.xlsx", cutoff_minutes=2):
    df = load_file(file_path)

    # Basic column safety checks (helps you spot a renamed column fast)
    required = ["Duration", "Started At", "CutOff Amount"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"These expected columns are missing from the file: {missing}")

    # --- Duration -> seconds, and Failed flag -----------------------------
    cutoff_seconds = cutoff_minutes * 60
    df["Duration_Seconds"] = df["Duration"].apply(parse_duration_to_seconds)
    df["Is_Failed"] = df["Duration_Seconds"] < cutoff_seconds

    # --- Dates for grouping -------------------------------------------------
    df["Started At_Parsed"] = pd.to_datetime(
        df["Started At"], format="%b %d, %Y %I:%M %p", errors="coerce"
    )
    # Fallback for any rows that don't match the expected format
    still_missing = df["Started At_Parsed"].isna() & df["Started At"].notna()
    if still_missing.any():
        df.loc[still_missing, "Started At_Parsed"] = pd.to_datetime(
            df.loc[still_missing, "Started At"], errors="coerce"
        )
    df["Month"] = df["Started At_Parsed"].dt.to_period("M").astype(str)
    df["Week"] = df["Started At_Parsed"].dt.strftime("%G-W%V")  # ISO year-week

    # --- CutOff Amount as a number -------------------------------------
    df["CutOff Amount_Num"] = pd.to_numeric(df["CutOff Amount"], errors="coerce")

    failed_df = df[df["Is_Failed"]].copy()

    # --- Monthly summary ------------------------------------------------
    monthly = (
        df.groupby("Month")
        .agg(
            Total_Sessions=("Transaction Id", "count"),
            Failed_Sessions=("Is_Failed", "sum"),
        )
        .reset_index()
    )
    monthly["Failed_%"] = (monthly["Failed_Sessions"] / monthly["Total_Sessions"] * 100).round(1)
    monthly["Success_%"] = (100 - monthly["Failed_%"]).round(1)
    failed_cutoff_by_month = (
        failed_df.groupby("Month")["CutOff Amount_Num"]
        .agg(Total_CutOff_Amount="sum", Avg_CutOff_Amount="mean")
        .reset_index()
    )
    monthly = monthly.merge(failed_cutoff_by_month, on="Month", how="left").fillna(0)
    monthly = monthly.sort_values("Month")
    monthly = monthly[["Month", "Total_Sessions", "Failed_Sessions", "Failed_%", "Success_%",
                        "Total_CutOff_Amount", "Avg_CutOff_Amount"]]

    # --- Weekly summary --------------------------------------------------
    weekly = (
        df.groupby("Week")
        .agg(
            Total_Sessions=("Transaction Id", "count"),
            Failed_Sessions=("Is_Failed", "sum"),
        )
        .reset_index()
    )
    weekly["Failed_%"] = (weekly["Failed_Sessions"] / weekly["Total_Sessions"] * 100).round(1)
    weekly["Success_%"] = (100 - weekly["Failed_%"]).round(1)
    failed_cutoff_by_week = (
        failed_df.groupby("Week")["CutOff Amount_Num"]
        .agg(Total_CutOff_Amount="sum", Avg_CutOff_Amount="mean")
        .reset_index()
    )
    weekly = weekly.merge(failed_cutoff_by_week, on="Week", how="left").fillna(0)
    weekly = weekly.sort_values("Week")
    weekly = weekly[["Week", "Total_Sessions", "Failed_Sessions", "Failed_%", "Success_%",
                      "Total_CutOff_Amount", "Avg_CutOff_Amount"]]

    # --- CutOff Amount band summary (only for failed sessions) -----------
    bins = [-1, 0, 50, 100, 250, 500, 1000, float("inf")]
    labels = ["0", "1-50", "51-100", "101-250", "251-500", "501-1000", "1000+"]
    failed_df["CutOff_Band"] = pd.cut(failed_df["CutOff Amount_Num"], bins=bins, labels=labels)
    cutoff_summary = (
        failed_df.groupby("CutOff_Band", observed=True)
        .agg(Failed_Sessions=("Transaction Id", "count"), Total_CutOff_Amount=("CutOff Amount_Num", "sum"))
        .reset_index()
    )

    # --- Write Excel report ------------------------------------------------
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.drop(columns=["Started At_Parsed"]).to_excel(writer, sheet_name="All Sessions", index=False)
        failed_df.drop(columns=["Started At_Parsed"]).to_excel(writer, sheet_name="Failed Sessions", index=False)
        monthly.to_excel(writer, sheet_name="Monthly Summary", index=False)
        weekly.to_excel(writer, sheet_name="Weekly Summary", index=False)
        cutoff_summary.to_excel(writer, sheet_name="CutOff Summary", index=False)

    # --- Highlight failed rows in "All Sessions" in red -------------------
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill

    wb = load_workbook(output_path)
    ws = wb["All Sessions"]
    fail_col_idx = list(df.drop(columns=["Started At_Parsed"]).columns).index("Is_Failed") + 1
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=fail_col_idx).value is True:
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = red_fill
    wb.save(output_path)

    # --- PDF report (one-page, styled like a formal statement) -----------
    pdf_path = str(Path(output_path).with_suffix(".pdf"))
    period_from = df["Started At_Parsed"].min()
    period_to = df["Started At_Parsed"].max()
    generate_pdf_report(df, monthly, weekly, cutoff_summary, cutoff_minutes,
                         pdf_path=pdf_path, period_from=period_from, period_to=period_to)

    # --- Console summary -------------------------------------------------
    total = len(df)
    failed = int(df["Is_Failed"].sum())
    success = total - failed
    print(f"Total sessions        : {total}")
    print(f"Failed sessions (<{cutoff_minutes}m)  : {failed} ({failed/total*100:.1f}%)")
    print(f"Success sessions       : {success} ({success/total*100:.1f}%)")
    print(f"Excel report saved to  : {output_path}")
    print(f"PDF report saved to    : {pdf_path}")

    return {
        "all_sessions": df,
        "failed_sessions": failed_df,
        "monthly": monthly,
        "weekly": weekly,
        "cutoff_summary": cutoff_summary,
    }


# --------------------------------------------------------------------------
# 3. Command-line entry point
# --------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Analyze failed EV charging sessions.")
    parser.add_argument("input_file", help="Path to the CSV or Excel export (e.g. ChargingSessions.csv)")
    parser.add_argument("--minutes", type=float, default=2, help="Failure cutoff in minutes (default: 2)")
    parser.add_argument("--output", default="Charging Session Report.xlsx", help="Output Excel file name")
    args = parser.parse_args()

    analyze(args.input_file, output_path=args.output, cutoff_minutes=args.minutes)


if __name__ == "__main__":
    main()
